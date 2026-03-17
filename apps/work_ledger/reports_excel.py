# =============================================================================
# FILE: apps/work_ledger/reports_excel.py
# Monthly Work Report Excel export for the active Work Ledger models.
# =============================================================================
import io
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional dependency
    OPENPYXL_AVAILABLE = False
else:
    OPENPYXL_AVAILABLE = True


MONTH_NAMES = [
    '', 'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]

NAVY_HEX = '003366'
GOLD_HEX = 'CC9900'
LIGHT_HEX = 'EEF2F7'
WHITE_HEX = 'FFFFFF'
GREY_HEX = '999999'


def _display_name(user) -> str:
    return (
        getattr(user, 'full_name', '')
        or getattr(user, 'get_full_name', lambda: '')()
        or str(user)
    )


def _fill(color: str):
    return PatternFill(fill_type='solid', fgColor=color)


def _font(*, bold: bool = False, color: str = '000000', size: int = 10):
    return Font(name='Calibri', size=size, bold=bold, color=color)


def _border():
    thin = Side(style='thin', color=GREY_HEX)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _left():
    return Alignment(horizontal='left', vertical='top', wrap_text=True)


def _get_entries(user, year: int, month: int):
    from apps.work_ledger.models import WorkEntry

    return list(
        WorkEntry.objects.filter(
            user=user,
            work_date__year=year,
            work_date__month=month,
        )
        .order_by('work_date', 'created_at')
        .select_related('category')
        .prefetch_related('verifications')
    )


def generate_monthly_work_report_excel(user, year: int, month: int) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl is required for Excel export. Run: pip install openpyxl')

    entries = _get_entries(user, year, month)
    month_label = f'{MONTH_NAMES[month]} {year}'
    full_name = _display_name(user)
    designation = getattr(user, 'designation', '') or ''
    section = getattr(user, 'section', '') or ''
    employee_id = getattr(user, 'employee_id', '') or ''

    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = 'Work Entries'
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells('A1:H1')
    sheet['A1'] = 'PATIALA LOCOMOTIVE WORKS - LOCO DRAWING OFFICE'
    sheet['A1'].font = _font(bold=True, color=WHITE_HEX, size=13)
    sheet['A1'].fill = _fill(NAVY_HEX)
    sheet['A1'].alignment = _center()

    sheet.merge_cells('A2:H2')
    sheet['A2'] = f'MONTHLY WORK REPORT - {month_label.upper()}'
    sheet['A2'].font = _font(bold=True, color=WHITE_HEX, size=11)
    sheet['A2'].fill = _fill(NAVY_HEX)
    sheet['A2'].alignment = _center()

    info_rows = [
        ('Staff Name', full_name, 'Employee ID', employee_id),
        ('Designation', designation, 'Section', section),
        ('Report Month', month_label, 'Generated On', date.today().strftime('%d-%b-%Y')),
    ]
    for row_index, (label_left, value_left, label_right, value_right) in enumerate(info_rows, start=3):
        sheet.cell(row_index, 1, label_left).font = _font(bold=True, color=NAVY_HEX)
        sheet.cell(row_index, 2, value_left)
        sheet.cell(row_index, 5, label_right).font = _font(bold=True, color=NAVY_HEX)
        sheet.cell(row_index, 6, value_right)

    headers = [
        'Sr',
        'Date',
        'Work Type',
        'Category',
        'Description',
        'Reference',
        'Effort',
        'Status',
    ]
    widths = [6, 14, 22, 18, 44, 20, 16, 16]
    header_row = 7

    for column_index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(header_row, column_index, header)
        cell.font = _font(bold=True, color=WHITE_HEX)
        cell.fill = _fill(NAVY_HEX)
        cell.alignment = _center()
        cell.border = _border()
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    total_effort = 0.0
    summary_counts: dict[str, int] = {}

    for entry_index, entry in enumerate(entries, start=1):
        row_index = header_row + entry_index
        effort_value = float(entry.effort_value or 0)
        total_effort += effort_value
        work_type_label = entry.get_work_type_display()
        summary_counts[work_type_label] = summary_counts.get(work_type_label, 0) + 1
        row_fill = _fill(LIGHT_HEX if entry_index % 2 == 0 else WHITE_HEX)
        values = [
            entry_index,
            entry.work_date.strftime('%d-%b-%Y') if entry.work_date else '',
            work_type_label,
            entry.category.name if entry.category else '',
            entry.description,
            entry.reference_number or '',
            '' if entry.effort_value is None else f'{entry.effort_value} {entry.get_effort_unit_display()}',
            entry.get_status_display(),
        ]

        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.fill = row_fill
            cell.border = _border()
            cell.alignment = _center() if column_index in (1, 2, 7, 8) else _left()
            cell.font = _font(size=9)

    total_row = header_row + len(entries) + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    total_label = sheet.cell(total_row, 1, 'TOTAL EFFORT')
    total_label.font = _font(bold=True, color=WHITE_HEX)
    total_label.fill = _fill(GOLD_HEX)
    total_label.alignment = _center()
    for column_index in range(1, 7):
        sheet.cell(total_row, column_index).fill = _fill(GOLD_HEX)
    total_value = sheet.cell(total_row, 7, round(total_effort, 2))
    total_value.font = _font(bold=True)
    total_value.fill = _fill(GOLD_HEX)
    total_value.alignment = _center()
    sheet.cell(total_row, 8).fill = _fill(GOLD_HEX)

    summary_sheet = workbook.create_sheet('Summary')
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.column_dimensions['A'].width = 34
    summary_sheet.column_dimensions['B'].width = 16

    summary_sheet.merge_cells('A1:B1')
    summary_sheet['A1'] = f'Work Type Summary - {month_label}'
    summary_sheet['A1'].font = _font(bold=True, color=WHITE_HEX, size=12)
    summary_sheet['A1'].fill = _fill(NAVY_HEX)
    summary_sheet['A1'].alignment = _center()

    for column_index, header in enumerate(['Work Type', 'Entry Count'], start=1):
        cell = summary_sheet.cell(2, column_index, header)
        cell.font = _font(bold=True, color=WHITE_HEX)
        cell.fill = _fill(NAVY_HEX)
        cell.alignment = _center()
        cell.border = _border()

    row_index = 3
    for work_type_label, count in sorted(summary_counts.items(), key=lambda item: (-item[1], item[0])):
        for column_index, value in enumerate([work_type_label, count], start=1):
            cell = summary_sheet.cell(row_index, column_index, value)
            cell.border = _border()
            cell.alignment = _center()
            cell.font = _font(size=9)
        row_index += 1

    for column_index, value in enumerate(['TOTAL ENTRIES', len(entries)], start=1):
        cell = summary_sheet.cell(row_index, column_index, value)
        cell.font = _font(bold=True)
        cell.fill = _fill(GOLD_HEX)
        cell.alignment = _center()
        cell.border = _border()

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
