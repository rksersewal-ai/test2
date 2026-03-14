# =============================================================================
# FILE: apps/pl_master/management/commands/import_category_book.py
#
# PURPOSE: Bulk-import Category Book data from Excel (.xlsx) or CSV into:
#          PLMaster, DrawingMaster, SpecificationMaster,
#          PLDrawingLink, PLSpecLink, PLStandardLink, PLLocoApplicability
#
# USAGE:
#   python manage.py import_category_book --file path/to/CategoryBook.xlsx
#   python manage.py import_category_book --file path/to/book.csv --sheet "WAG9"
#   python manage.py import_category_book --file book.xlsx --dry-run
#   python manage.py import_category_book --file book.xlsx --skip-errors --agency CLW
#
# EXPECTED COLUMN HEADERS (case-insensitive, spaces/underscores interchangeable):
#   pl_number, part_description, part_description_hindi,
#   drawing_numbers, spec_number, material_spec,
#   applicable_standards, unit_of_measure, inspection_category,
#   stage_inspection, inspection_agency, controlling_agency,
#   vendor_registration_category, smi_reference,
#   interchangeability_code, safety_classification, shelf_life,
#   last_amendment_number, remarks
#   (plus optional: used_in, uvam_id, mother_part, vd_item, nvd_item)
# =============================================================================
import os
import re
import csv
import json
import traceback
from datetime import date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Column name normaliser  (maps many possible header spellings to canonical)
# ---------------------------------------------------------------------------
COL_MAP = {
    # PL Master core
    'pl_number':                    'pl_number',
    'pl number':                    'pl_number',
    'item_code':                    'pl_number',
    'item code':                    'pl_number',
    'part_description':             'part_description',
    'part description':             'part_description',
    'description':                  'part_description',
    'item_description':             'part_description',
    'item description':             'part_description',
    'part_description_hindi':       'part_description_hindi',
    'description_hindi':            'part_description_hindi',
    'hindi description':            'part_description_hindi',
    # Drawing
    'drawing_numbers':              'drawing_numbers',
    'drawing numbers':              'drawing_numbers',
    'drawing_no':                   'drawing_numbers',
    'drg no':                       'drawing_numbers',
    # Spec
    'spec_number':                  'spec_number',
    'specification_number':         'spec_number',
    'spec number':                  'spec_number',
    'material_spec':                'material_spec',
    'material spec':                'material_spec',
    # Standards
    'applicable_standards':         'applicable_standards',
    'standards':                    'applicable_standards',
    'applicable standards':         'applicable_standards',
    # UOM
    'unit_of_measure':              'unit_of_measure',
    'uom':                          'unit_of_measure',
    'unit':                         'unit_of_measure',
    # Inspection
    'inspection_category':          'inspection_category',
    'category':                     'inspection_category',
    'stage_inspection':             'stage_inspection',
    'stage_inspection_reqd':        'stage_inspection',
    'inspection_agency':            'inspection_agency',
    'inspection agency':            'inspection_agency',
    # Controlling
    'controlling_agency':           'controlling_agency',
    'controlling authority':        'controlling_agency',
    # Vendor / UVAM
    'vendor_registration_category': 'vendor_registration_category',
    'vendor category':              'vendor_registration_category',
    'vd_item':                      'vd_item',
    'vd item':                      'vd_item',
    'nvd_item':                     'nvd_item',
    'uvam_id':                      'uvam_id',
    'uvam id':                      'uvam_id',
    # SMI
    'smi_reference':                'smi_reference',
    'smi reference':                'smi_reference',
    'smi':                          'smi_reference',
    # Interchangeability
    'interchangeability_code':      'interchangeability_code',
    'interchangeability code':      'interchangeability_code',
    # Safety
    'safety_classification':        'safety_classification',
    'criticality_classification':   'safety_classification',
    'criticality':                  'safety_classification',
    # Misc
    'shelf_life':                   'shelf_life',
    'last_amendment_number':        'last_amendment_number',
    'last amendment':               'last_amendment_number',
    'remarks':                      'remarks',
    # Extended optional
    'used_in':                      'used_in',
    'loco_type':                    'used_in',
    'loco type':                    'used_in',
    'mother_part':                  'mother_part',
    'mother part':                  'mother_part',
    'parent_pl':                    'mother_part',
}


def normalise_col(raw_header):
    """Map raw header string to canonical field name or return None."""
    return COL_MAP.get(raw_header.strip().lower().replace('-', '_'))


def str_val(cell_value):
    """Return stripped string or empty string."""
    if cell_value is None:
        return ''
    return str(cell_value).strip()


def parse_list(val, sep=','):
    """Split a delimited cell into a cleaned list of non-empty strings."""
    if not val:
        return []
    return [x.strip() for x in re.split(r'[,;/|]', val) if x.strip()]


def parse_bool(val):
    return str_val(val).upper() in ('YES', 'Y', '1', 'TRUE')


def map_inspection_category(val):
    mapping = {
        'A': 'CAT-A', 'CAT-A': 'CAT-A', 'CAT A': 'CAT-A',
        'B': 'CAT-B', 'CAT-B': 'CAT-B', 'CAT B': 'CAT-B',
        'C': 'CAT-C', 'CAT-C': 'CAT-C', 'CAT C': 'CAT-C',
        'D': 'CAT-D', 'CAT-D': 'CAT-D', 'CAT D': 'CAT-D',
    }
    return mapping.get(str_val(val).upper(), 'CAT-D')


def map_safety(val):
    v = str_val(val).upper()
    if 'CRITICAL' in v:
        return 'CRITICAL'
    if 'HIGH' in v or 'SAFETY' in v:
        return 'HIGH'
    if 'MEDIUM' in v or 'MOD' in v:
        return 'MEDIUM'
    return 'LOW'


def map_insp_agency(val):
    v = str_val(val).upper()
    if 'RITES' in v:
        return 'RITES'
    if 'THIRD' in v or 'TPI' in v:
        return 'THIRD_PARTY'
    return 'SELF'


# ---------------------------------------------------------------------------
class Command(BaseCommand):
    help = 'Bulk-import Category Book data from Excel/CSV into PLMaster'

    def add_arguments(self, parser):
        parser.add_argument('--file',        required=True, help='Path to .xlsx or .csv file')
        parser.add_argument('--sheet',       default=None,  help='Sheet name (xlsx only). Omit to process ALL sheets.')
        parser.add_argument('--agency',      default=None,  help='Override controlling_agency for all rows (e.g. CLW)')
        parser.add_argument('--dry-run',     action='store_true', help='Parse and validate only, do not write to DB')
        parser.add_argument('--skip-errors', action='store_true', help='Continue on row-level errors')
        parser.add_argument('--update',      action='store_true', help='Update existing PLMaster records (default: skip)')

    def handle(self, *args, **options):
        filepath    = options['file']
        sheet_name  = options['sheet']
        agency_ovr  = options['agency']
        dry_run     = options['dry_run']
        skip_errors = options['skip_errors']
        do_update   = options['update']

        if not os.path.exists(filepath):
            raise CommandError(f'File not found: {filepath}')

        ext = os.path.splitext(filepath)[1].lower()

        if ext in ('.xlsx', '.xls'):
            if not OPENPYXL_AVAILABLE:
                raise CommandError('openpyxl is required for Excel import. Run: pip install openpyxl')
            sheets = self._load_excel_sheets(filepath, sheet_name)
        elif ext == '.csv':
            sheets = {'CSV': self._load_csv(filepath)}
        else:
            raise CommandError(f'Unsupported file type: {ext}. Use .xlsx or .csv')

        total_stats = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        for sname, rows in sheets.items():
            self.stdout.write(self.style.HTTP_INFO(f'\n=== Sheet: {sname} ({len(rows)} rows) ==='))
            stats = self._process_rows(
                rows, agency_ovr, dry_run, skip_errors, do_update
            )
            for k in total_stats:
                total_stats[k] += stats[k]

        self.stdout.write('\n' + '='*60)
        mode = '[DRY-RUN] ' if dry_run else ''
        self.stdout.write(self.style.SUCCESS(
            f'{mode}Import complete:\n'
            f'  Processed : {total_stats["processed"]}\n'
            f'  Created   : {total_stats["created"]}\n'
            f'  Updated   : {total_stats["updated"]}\n'
            f'  Skipped   : {total_stats["skipped"]}\n'
            f'  Errors    : {total_stats["errors"]}'
        ))

    # -----------------------------------------------------------------------
    def _load_excel_sheets(self, filepath, sheet_filter):
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = {}
        names  = [sheet_filter] if sheet_filter else wb.sheetnames
        for sname in names:
            if sname not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sname}" not found, skipping.'))
                continue
            ws   = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            header_raw = [str(c) if c is not None else '' for c in rows[0]]
            header     = [normalise_col(h) for h in header_raw]
            data_rows  = []
            for row in rows[1:]:
                if all(c is None or str(c).strip() == '' for c in row):
                    continue  # skip fully blank rows
                data_rows.append(dict(zip(header, [str_val(c) for c in row])))
            sheets[sname] = data_rows
        wb.close()
        return sheets

    def _load_csv(self, filepath):
        rows = []
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            header = {col: normalise_col(col) for col in reader.fieldnames or []}
            for raw_row in reader:
                row = {header[k]: str_val(v) for k, v in raw_row.items() if header.get(k)}
                rows.append(row)
        return rows

    # -----------------------------------------------------------------------
    def _process_rows(self, rows, agency_ovr, dry_run, skip_errors, do_update):
        from apps.pl_master.models import (
            PLMaster, DrawingMaster, SpecificationMaster,
            ControllingAgency, PLDrawingLink, PLSpecLink,
            PLStandardLink, PLLocoApplicability,
        )

        stats = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

        for idx, row in enumerate(rows, start=2):  # row 1 = header
            pl_num = row.get('pl_number', '').strip()
            if not pl_num:
                stats['skipped'] += 1
                continue

            stats['processed'] += 1
            try:
                with transaction.atomic():
                    self._upsert_row(
                        row, pl_num, agency_ovr, do_update,
                        dry_run, stats,
                    )
            except Exception as exc:
                stats['errors'] += 1
                msg = f'Row {idx} PL={pl_num}: {exc}'
                if skip_errors:
                    self.stdout.write(self.style.ERROR(f'  ERROR (skipped): {msg}'))
                else:
                    raise CommandError(msg) from exc

        return stats

    # -----------------------------------------------------------------------
    def _upsert_row(self, row, pl_num, agency_ovr, do_update, dry_run, stats):
        from apps.pl_master.models import (
            PLMaster, DrawingMaster, SpecificationMaster,
            ControllingAgency, PLDrawingLink, PLSpecLink,
            PLStandardLink, PLLocoApplicability,
        )

        # --- Resolve controlling agency ------------------------------------
        agency_code = agency_ovr or row.get('controlling_agency', 'CLW')
        agency_obj  = ControllingAgency.objects.filter(
            agency_code__iexact=agency_code
        ).first()

        # --- Build PLMaster field dict -------------------------------------
        insp_cat    = map_inspection_category(row.get('inspection_category', ''))
        safety_raw  = row.get('safety_classification', '')
        is_safety   = parse_bool(safety_raw) or ('CRITICAL' in safety_raw.upper() or 'HIGH' in safety_raw.upper())
        safety_cls  = map_safety(safety_raw) if is_safety else 'LOW'

        used_in_raw = row.get('used_in', '')
        used_in     = parse_list(used_in_raw) if used_in_raw else []

        stage_raw   = row.get('stage_inspection', '')
        stage_json  = self._parse_stage_inspection(stage_raw)

        pl_fields = dict(
            part_description        = row.get('part_description', '')[:500] or f'PL {pl_num}',
            part_description_hindi  = row.get('part_description_hindi', ''),
            inspection_category     = insp_cat,
            safety_item             = is_safety,
            safety_classification   = safety_cls,
            inspection_agency       = map_insp_agency(row.get('inspection_agency', '')),
            stage_inspection_reqd   = stage_json,
            vd_item                 = parse_bool(row.get('vd_item', '')),
            nvd_item                = parse_bool(row.get('nvd_item', '')),
            uvam_id                 = row.get('uvam_id', ''),
            used_in                 = used_in,
            remarks                 = row.get('remarks', ''),
            is_active               = True,
        )
        if agency_obj:
            pl_fields['controlling_agency'] = agency_obj

        mother_pl = row.get('mother_part', '').strip()
        # Mother part FK resolved after all rows ideally, skip for now

        # --- Create or update PLMaster ------------------------------------
        existing = PLMaster.objects.filter(pl_number=pl_num).first()
        if existing:
            if do_update:
                if not dry_run:
                    for f, v in pl_fields.items():
                        setattr(existing, f, v)
                    existing.save()
                stats['updated'] += 1
                self.stdout.write(f'  UPDATED  {pl_num}')
            else:
                stats['skipped'] += 1
                return
        else:
            if not dry_run:
                pl_obj = PLMaster.objects.create(pl_number=pl_num, **pl_fields)
            else:
                pl_obj = PLMaster(pl_number=pl_num, **pl_fields)  # unsaved
            stats['created'] += 1
            self.stdout.write(f'  CREATED  {pl_num}')

        if dry_run:
            return  # Do not create any linked records in dry-run

        # Re-fetch object for linking (handles create path)
        pl_obj = PLMaster.objects.get(pl_number=pl_num)

        # --- Drawings -------------------------------------------------------
        drg_raw  = row.get('drawing_numbers', '')
        drawings = parse_list(drg_raw)
        for drg_num in drawings:
            drg_obj, _ = DrawingMaster.objects.get_or_create(
                drawing_number=drg_num,
                defaults={
                    'drawing_title': f'Drawing {drg_num}',
                    'drawing_type':  'CD',
                    'controlling_agency': agency_code,
                    'current_alteration': '00',
                    'is_active': True,
                }
            )
            PLDrawingLink.objects.get_or_create(
                pl_master=pl_obj, drawing=drg_obj,
                defaults={'is_primary': drawings.index(drg_num) == 0}
            )

        # --- Specification --------------------------------------------------
        spec_num = row.get('spec_number', '').strip()
        if spec_num:
            spec_obj, _ = SpecificationMaster.objects.get_or_create(
                spec_number=spec_num,
                defaults={
                    'spec_title': f'Specification {spec_num}',
                    'spec_type':  'TS',
                    'controlling_agency': agency_code,
                    'current_alteration': '00',
                    'is_active': True,
                }
            )
            PLSpecLink.objects.get_or_create(
                pl_master=pl_obj, specification=spec_obj
            )

        # Material spec treated as additional specification
        mat_spec = row.get('material_spec', '').strip()
        if mat_spec and mat_spec != spec_num:
            mspec_obj, _ = SpecificationMaster.objects.get_or_create(
                spec_number=mat_spec,
                defaults={
                    'spec_title': f'Material Spec {mat_spec}',
                    'spec_type':  'MS',
                    'controlling_agency': agency_code,
                    'current_alteration': '00',
                    'is_active': True,
                }
            )
            PLSpecLink.objects.get_or_create(
                pl_master=pl_obj, specification=mspec_obj
            )

        # --- Standards ------------------------------------------------------
        stds = parse_list(row.get('applicable_standards', ''))
        from apps.pl_master.models import RDSOReference
        for std_num in stds:
            std_obj, _ = RDSOReference.objects.get_or_create(
                rdso_doc_number=std_num,
                defaults={
                    'rdso_doc_title': f'Standard {std_num}',
                    'rdso_doc_type':  'SPEC',
                    'is_active': True,
                }
            )
            PLStandardLink.objects.get_or_create(
                pl_master=pl_obj, standard=std_obj
            )

        # --- Loco Applicability ---------------------------------------------
        for loco in (used_in or []):
            PLLocoApplicability.objects.get_or_create(
                pl_master=pl_obj, loco_type=loco
            )

    # -----------------------------------------------------------------------
    def _parse_stage_inspection(self, raw):
        """Map free-text stage inspection cell to structured JSONB dict."""
        v    = raw.upper()
        data = {
            'RMS': 'RMS' in v or 'RAW MATERIAL' in v,
            'PPS': 'PPS' in v or 'PRE-PROCESS' in v or 'PREPROCESS' in v,
            'IPS': 'IPS' in v or 'IN-PROCESS' in v or 'INPROCESS' in v,
            'FMS': 'FMS' in v or 'FINAL MACHINE' in v,
            'HTS': 'HTS' in v or 'HEAT TREAT' in v,
            'FIS': 'FIS' in v or 'FINAL INSP' in v or bool(v),  # FIS always True if any stage specified
        }
        return data
